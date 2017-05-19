#!/usr/bin/env python
# Copyright (C) 2015,2016 Mohammad Alanjary
# University of Tuebingen
# Interfaculty Institute of Microbiology and Infection Medicine
# Lab of Nadine Ziemert, Div. of Microbiology/Biotechnology
# Funding by the German Centre for Infection Research (DZIF)
#
# This file is part of ARTS
# ARTS is free software. you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version
#
# License: You should have received a copy of the GNU General Public License v3 with ARTS
# A copy of the GPLv3 can also be found at: <http://www.gnu.org/licenses/>.

import os
import urllib2
import tempfile
import re
import time
import json
#import subprocess

from openpyxl import Workbook
#from ete3 import Tree
from flask import request, g, flash
from flask_mail import Message
from redis import Redis
from redis import ConnectionError as redisConnectError
from werkzeug.utils import secure_filename
from Bio import Entrez

from app import app
from app import mail
from app import models


#load model metadata
# modeldata = {}
# if os.path.exists(os.path.join(app.config["REF_FOLDER"],"model_metadata.tsv")):
#     with open(os.path.join(app.config["REF_FOLDER"],"model_metadata.tsv")) as fil:
#         for line in fil:
#             if not line.startswith("#"):
#                 x = line.strip().split("\t")
#                 if x[0] not in modeldata:
#                     modeldata[x[0]] = x[1:]

def sendnotifymail(msg="",jobid="",to=""):
    try:
        if not msg:
            msg = "Hello, your ARTS job has been submitted! Your job id is: "
        assert to, jobid
        msgobj = Message("Your ARTS Job (%s) has been submitted"%jobid,recipients=[to])
        msgobj.html = "%s %s <br> <a href='https://arts.ziemertlab.com/results/%s'>https://arts.ziemertlab.com/results/%s</a>"%(msg,jobid,jobid,jobid)
        mail.send(msgobj)
    except Exception as e:
        print "Warning: Email not sent, check email configuration"
        print e

def which(program):
    def is_exe(fpath):
        return os.path.isfile(fpath) and os.access(fpath, os.X_OK)

    fpath, fname = os.path.split(program)
    if fpath:
        if is_exe(program):
            return program
    else:
        for path in os.environ["PATH"].split(os.pathsep):
            path = path.strip('"')
            exe_file = os.path.join(path, program)
            if is_exe(exe_file):
                return exe_file
    return None

# def getdupmatrix(rd):
#     with open(os.path.join(rd,"resultsummary.json"),"r") as fil:
#         rs = json.load(fil)
#         dupmat = [[x]+rs["matrix"][x][:5] for x in rs["duplicates"]]
#         return {"data":dupmat}
#
# def getbgctable(rd):
#     with open(os.path.join(rd,"resultsummary.json"),"r") as fil:
#         rs = json.load(fil)
#         proxhits = rs["proximity"]
#         temp = {}
#         for hk in proxhits.keys():
#             hitdict = proxhits[hk]
#             for seqid,row in hitdict.items():
#                 clustid = "cluster-%s"%row[0]
#                 if clustid not in temp:
#                     temp[clustid] = {"row":[clustid,row[1],row[2],"%s - %s"%(row[3],row[4])],"hits":[]}
#                 temp[clustid]["hits"].append([seqid,hk,str(row[5]),str(row[6]),modeldata[hk][0]+": "+modeldata[hk][1],modeldata[hk][2]])
#         data=[]
#         for x in temp.values():
#             data.append(x["row"]+[str(len(x["hits"]))]+[x["hits"]])
#         temp["data"] = data
#         return temp

# def rendertree(infile, width=800, spname=False):
#     #First try workaround for no Xserver
#     T = Tree(infile)
#     if which("xvfb-run"):
#         rtfile = os.path.join(os.path.split(os.path.realpath(__file__))[0],"rendertree.py")
#         child = subprocess.Popen(["xvfb-run","python",str(rtfile),"-in",str(infile),"-w",str(width),"-spname",str(spname)], stdout=subprocess.PIPE)
#         output = child.communicate()[0]
#         if child.returncode==0:
#             return True
#         return False
#     elif len(app.config.get("DISABLE_PYQT",False)):
#         printpng(str(T),spname,"fonts/roboto.tff",16,offset=10)
#     else:
#         if spname:
#             try:
#                 for node in T:
#                     if "OUTGROUP" in node.name:
#                         T.set_outgroup(node)
#                         T.ladderize()
#                     if spname in node.name:
#                         #Use styles if availible (PyQT) otherwise use TEXT
#                         from ete3 import NodeStyle
#                         nstyle = NodeStyle()
#                         nstyle["bgcolor"] = "#00CC00"
#                         nstyle["size"] = 10
#                         node.set_style(nstyle)
#                 return T.render(infile+".png", w=width, units="px")
#             except ImportError:
#                 printpng(str(T),spname,"fonts/roboto.tff",16,offset=10)

def getdb():
    rddb = getattr(g,"_redisdb",False)
    if not rddb:
        rddb = g._redisdb = Redis.from_url(app.config['REDISURL'])
    try:
        rddb.ping()
    except redisConnectError:
        rddb = False
    return rddb

def validatehmm(fname):
    validext = ['hmm']
    ext = os.path.splitext(fname)[1]
    if ext[1:].lower() in validext:
        return True
    return False

def validatefile(fname, asfil=False):
    validgbkext = ['gbk','genbank','gbff','gb','embl']
    validfakext = ['fasta','fa','fna','faa','fas']
    ext = os.path.splitext(fname)[1]
    if not asfil and ext[1:].lower() in validgbkext+validfakext:
        return True
    elif asfil and ext[1:].lower() in validgbkext:
        return True
    return False

def getoptions():
    options = []
    x = request.form.get('run-phyl',False)
    if x and x.lower() != "false":
        options.append("phyl")
    x = request.form.get('run-kres',False)
    if x and x.lower() != "false":
        options.append("kres")
    x = request.form.get('run-duf',False)
    if x and x.lower() != "false":
        options.append("duf")
    x = request.form.get('run-expert',False)
    if x and x.lower() != "false":
        options.append("expert")
    return ",".join(options)

def getNCBIgbk(acc):
    try:
        #crude check for format
        if acc.upper().startswith("GCA_") or acc.upper().startswith("GCF_"):
            flash("Error, cannot use assembly accession number, please use Genbank or Refseq sequence accession")
            filename = False
        elif acc.replace("_","").replace(".","").isalnum() and len(acc) <= 20 and len(acc) >= 6:
            if "." in acc:
                acc = acc[:acc.index(".")]
            Entrez.email = "artsuser@ziemertlab.com"
            handle = Entrez.efetch(db="nucleotide", rettype="gb", id=acc, retmode="text")
            filename = os.path.join(tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER']), secure_filename(acc+".gbk"))
            with open(filename,"w") as outfile:
                outfile.write(handle.read())
        else:
            flash("Error with Accession number ")
            filename = False
    except Exception, e:
        flash("Error retrieving gbk from NCBI")
        filename = False
    return filename

def getASgbk(asjobid):
    try:
        temp = urllib2.urlopen(os.path.join(app.config["ASMSHSERVER"],"upload",asjobid,"index.html"))
        filename=""
        for line in temp.readlines():
            m = re.search('href=[\'"]?([^\'" >]+)\.final\.gbk',line)
            if m:
                filename = m.group(1)+".final.gbk"
                aslink = os.path.join(app.config["ASMSHSERVER"],"upload",asjobid,filename)
                #Save file
                filename = os.path.join(tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER']), secure_filename(filename))
                dlfile = urllib2.urlopen(aslink)
                with open(filename,'wb') as outfile:
                    outfile.write(dlfile.read())
                break
    except urllib2.HTTPError, e:
        flash("Failed to get antismash genbank result from job id")
        filename = False
    except urllib2.URLError, e:
        flash("Failed to get antismash genbank result from job id")
        filename = False
    return filename

def getASstatus(asjobid):
    #quick validation
    if asjobid:
        try:
            temp = urllib2.urlopen(os.path.join(app.config["ASMSHSERVER"],"upload",asjobid,"index.html"))
            return True
        except urllib2.HTTPError:
            return False
    return False

def getcustmdl(reqid):
    if request.files.get(reqid,False) and validatehmm(request.files[reqid].filename):
        ufile = request.files[reqid]
        filename = os.path.join(tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER']), secure_filename(ufile.filename))
        ufile.save(filename)
        return filename
    elif request.files.get(reqid,False):
        flash("Failed to load custom models, please check format and try again")
        return "Failed"
    return False

def getinfile():
    ufile = False
    filename = ""
    # USe empty string for FALSE due to storage as string in redis
    asrun = ""
    if 'asjobid' in request.form and request.form['asjobid']:
        if getASstatus(request.form['asjobid']):
            filename = getASgbk(request.form['asjobid'])
            if not filename:
                return False,""
        else:
            return False,""
    elif 'ncbiacc' in request.form and request.form['ncbiacc']:
        filename = getNCBIgbk(request.form['ncbiacc'])
        asrun = True
        if not filename:
            return False,""
    elif 'asseqfile' in request.files and validatefile(request.files['asseqfile'].filename,True):
        ufile = request.files['asseqfile']
    elif 'seqfile' in request.files and validatefile(request.files['seqfile'].filename):
        ufile = request.files['seqfile']
        asrun = True
    if ufile:
        filename = os.path.join(tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER']), secure_filename(ufile.filename))
        ufile.save(filename)
    return filename,asrun

def addjob(**kwargs):
    rddb = getdb()
    artsjob = models.ArtsJob(**kwargs)
    if rddb:
        rddb.hmset("artsjob:%s"%artsjob.id,artsjob.getdict())
        if rddb.llen("PQ"):
            rddb.lpush("PQ",artsjob.id) #Store in the Pause Queue if paused
        else:
            rddb.lpush("SQ",artsjob.id) #Store in the Start Queue
        os.mkdir(os.path.join(app.config['RESULTS_FOLDER'],artsjob.id))
    return artsjob

# def formatkrtab(jobid,outfile=None):
#     kr = os.path.join(app.config['RESULTS_FOLDER'],jobid,"knownhits.json")
#     if os.path.exists(kr):
#         krlist = []
#         with open(kr,"r") as fil:
#             kr = json.load(fil)
#         for rec in kr.values():
#             for row in rec.values():
#                 krlist.append(row)
#         return {"data":krlist}
#     return {"data":[]}

# def formatsumtab(jobid,outfile=None):
#     rd = os.path.join(app.config['RESULTS_FOLDER'],jobid)
#     gm = os.path.join(rd,"resultsummary.json")
#     kr = os.path.join(rd,"knownhits.json")
#     krlist = {}
#     corelist = os.path.join(rd,"coregenes","coreresults.json")
#     if os.path.exists(gm) and os.path.exists(corelist):
#         with open(gm,"r") as fil:
#             gm = json.load(fil)
#         with open(corelist,"r") as fil:
#             corelist = json.load(fil)
#         if os.path.exists(kr):
#             with open(kr,"r") as fil:
#                 krlist = json.load(fil)
#         summary = {"data":[],"seqs":corelist["seqs"]}
#         funcstats = {}
#         for k in gm["matrix"].keys():
#             metadata = modeldata.get(k,["N/A","N/A",""])
#             if metadata[2]:
#                 if metadata[2] not in funcstats:
#                     funcstats[metadata[2]] = 0
#                 funcstats[metadata[2]] += 1
#             temp = {"coregene":k,"description":"%s: %s"%(metadata[0],metadata[1]),"func":metadata[2],"duplicate":"N/A","proximity":"N/A",
#                     "phylogeny":"N/A","hits":corelist["core"].get(k,{}).get("seqs",[]),"proxhits":[],"known_hit":"N/A"}
#             if "duplicates" in gm.keys():
#                 if k in gm["duplicates"]:
#                     temp["duplicate"] = "Yes"
#                 else:
#                     temp["duplicate"] = "No"
#             if "proximity" in gm.keys():
#                 if k in gm["proximity"].keys():
#                     temp["proximity"] = "Yes"
#                     temp["proxhits"] = gm["proximity"][k]
#                 else:
#                     temp["proximity"] = "No"
#             if "phylogeny" in gm.keys():
#                 if k in gm["phylogeny"].keys():
#                     temp["phylogeny"] = "Yes"
#                     temp["phylhits"] = gm["phylogeny"][k]
#                 else:
#                     temp["phylogeny"] = "No"
#             temp["known_hit"] = "No"
#             for seqid in temp["hits"]:
#                 if "seqs" in krlist and seqid in krlist["seqs"].keys():
#                     temp["known_hit"] = "Yes"
#             summary["data"].append(temp)
#         summary["funcstats"]=funcstats
#         if outfile:
#             with open(outfile,"w") as fil:
#                 json.dump(summary,fil,indent=2)
#         return summary
#     else:
#         return {"data":[]}


def getjobstatus(jobid):
    rddb = getdb()
    rdir = os.path.join(app.config['RESULTS_FOLDER'],jobid)
    log = os.path.join(rdir,"arts-query.log")
    aslog = os.path.join(rdir,"antismash","statusfile.txt")
    jobtitle = os.path.join(rdir,"jobtitle.txt")
    status = {
        "id":str(jobid),
        "state":"",
        "start":"",
        "end":"",
        "orgname":str(jobid),
        "jobtitle":str(jobid),
        "step":"",
        "tsteps":5,
        "buildtree":0,
        "coretotal":"N/A",
        "cdscount":"N/A",
        "bgccount":"N/A",
        "dupcount":"N/A",
        "phylcount":"N/A",
        "proxcount":"N/A",
        "twocount":"N/A",
        "threecount":"N/A",
        "krhits":"N/A",
    }
    if os.path.exists(rdir):
        ## Get state
        pwidth = "10%"
        ptitle = "Starting..."
        status["state"] = "Waiting in queue"
        if rddb and "artsjob:%s"%jobid in rddb.keys():
            strt,fin,err = rddb.hmget("artsjob:%s"%jobid,"started","finished","error")
            if fin:
                status["state"] = "Done"
                status["start"] = int(strt)
                status["end"] = int(fin)
            elif err:
                status["state"] = "Error"
                status["start"] = int(strt)
                status["end"] = int(err)
            elif strt:
                status["state"] = "Running"
                status["start"] = int(strt)
        if os.path.exists(aslog):
            with open(aslog,"r") as fil:
                status["state"] = "antiSMASH - %s"%fil.next().strip().replace("running:","")
        if os.path.exists(log):
            bt = 0
            tt = 0
            errors = 0
            warnings = 0
            #Parse Log
            with open(log, "r") as fil:
                for line in fil:
                    if "ERROR" in line:
                        errors+=1
                    if "WARNING" in line:
                        warnings+=1
                    if "query: org=" in line:
                        status["orgname"] = line.strip().split("org=")[-1]
                    if "CDS features:" in line:
                        temp = line.strip().split(";")
                        status["cdscount"] = int(temp[-2].split()[-1])
                        status["bgccount"] = int(temp[-1].split()[-1])
                    if "SUCCESS!" in line:
                        status["state"] = "Done"
                    if "duplicate genes" in line:
                        status["dupcount"] = line.split()[-3]
                    if "Proximity hits found" in line:
                        status["proxcount"] = line.split()[-1]
                    if "Hits with two or more criteria" in line:
                        status["twocount"] = line.split(":")[-2].strip()
                    if "Hits with three or more criteria" in line:
                        status["threecount"] = line.split(":")[-2].strip()
                    if "Known Resistance Hits:" in line:
                        status["krhits"] = line.split()[-1]
                    if "Phylogeny hits found:" in line:
                        status["phylcount"] = line.split()[-1]
                    if "Milestone_" in line:
                        status["step"] = line.strip().split("Milestone_")[-1][:1]
                    if "BuildTree: Finished" in line:
                        bt += 1
                    if "Wrote (1 of " in line:
                        tt = line.split("Wrote (1 of ")[1].split(")")[0]
            if tt:
                status["buildtree"] = int(100*float(bt)/float(tt))
                status["coretotal"] = int(tt)
                pwidth = str(30+status["buildtree"]/2)+"%"
                ptitle=pwidth + " - Step 2/5 done, Building trees..."
            if status["step"] ==  "1":
                pwidth="20%"
                ptitle=pwidth + " - Step 1/5 done, Extracting core genes..."
            if status["step"] ==  "3":
                pwidth="80%"
                ptitle=pwidth + " - Step 3/5 done, Building Species tree..."
            if status["step"] ==  "4":
                pwidth="90%"
                ptitle=pwidth + " - Step 4/5 done, Comparing trees..."
            if status["state"] ==  "Done":
                pwidth="100%"
                ptitle=pwidth + " Complete"
                if errors or warnings:
                    ptitle+=" with %s errors and %s warnings"%(errors,warnings)
            status["pwidth"] = pwidth
            status["ptitle"] = ptitle
        if os.path.exists(jobtitle):
            with open(jobtitle,"r") as fil:
                status["jobtitle"] = fil.next().strip()
        else:
            status["jobtitle"] = status["orgname"]
    return status

def checkresult(resid):
    rddb = getdb()
    if os.path.isdir(os.path.join(app.config["RESULTS_FOLDER"],resid)):
        return True
    elif rddb and "artsjob:%s"%resid in rddb.keys():
        return True
    return False

def getservstats():
    rddb = getdb()
    status={}
    if rddb:
        status["running"] = rddb.llen("RQ")
        status["waiting"] = rddb.llen("SQ")
        status["finished"] = rddb.llen("DQ")
        status["paused"] = rddb.llen("PQ")
    return status

#Read organism name from log file
def getqorg(rdir):
    log = os.path.join(rdir,"arts-query.log")
    qorg = False
    if os.path.exists(log):
        with open(log,"r") as fil:
            for line in fil:
                if "query: org=" in line:
                        qorg = line.strip().split("org=")[-1]
                        break
    return qorg

def getjobinfo(jobid):
    jobtitle = ""
    temp = {"id":jobid,"age":"","orgname":jobid,"jobtitle":jobtitle}
    rd = os.path.join(app.config["RESULTS_FOLDER"],jobid)
    if os.path.isdir(rd):
        orgname = getqorg(rd)
        orgname = orgname if orgname else jobid
        if os.path.exists(os.path.join(rd,"jobtitle.txt")):
            with open(os.path.join(rd,"jobtitle.txt"),"r") as fil:
                jobtitle = fil.next().strip()
        temp = {"id":jobid,"age":int(time.time()) - int(os.path.getctime(rd)),"orgname":orgname,"jobtitle":jobtitle}
    return temp

def getallresults():
    rlist = os.listdir(app.config["RESULTS_FOLDER"])
    return [getjobinfo(x) for x in rlist if checkresult(x) and x.lower() != "example"]

def getlastresults():
    lastresult = request.cookies.get('arts.lastresult',False)
    if lastresult:
        temp = lastresult.split(";")
        temp.reverse()
        lastresult = [getjobinfo(x) for x in temp]
    return lastresult

def makexltable(tpath):
    if not os.path.exists(os.path.join(tpath,"alltables.xlsx")):
        wb = Workbook()
        cws = wb.active
        cws.title = "Summary core hits"
        with open(os.path.join(tpath,"coretable.tsv"),"r") as infil:
            for line in infil:
                cws.append(line.strip().split("\t"))
        if os.path.exists(os.path.join(tpath,"bgctable.tsv")):
            bgcws = wb.create_sheet("BGC Proximity")
            with open(os.path.join(tpath,"bgctable.tsv"),"r") as infil:
                for line in infil:
                    row = line.strip().split("\t")
                    row[-1] = row[-1].replace("u'","").replace("'","")
                    bgcws.append(row)
        if os.path.exists(os.path.join(tpath,"duptable.tsv")):
            dupws = wb.create_sheet("BGC Proximity")
            with open(os.path.join(tpath,"duptable.tsv"),"r") as infil:
                for line in infil:
                    dupws.append(line.strip().split("\t"))
        if os.path.exists(os.path.join(tpath,"knownhits.tsv")):
            khws = wb.create_sheet("Resistance models")
            with open(os.path.join(tpath,"knownhits.tsv"),"r") as infil:
                for line in infil:
                    khws.append(line.strip().split("\t"))
        wb.save(os.path.join(tpath,"alltables.xlsx"))
    return "alltables.xlsx"

def json2dict(fpath):
    with open(fpath,"r") as fil:
        return json.load(fil)
    return {"error":"No such resource"}
